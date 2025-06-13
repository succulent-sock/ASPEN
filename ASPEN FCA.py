import subprocess
import sys
import time
import tkinter
import tkinter.messagebox
import tkinter.filedialog
import re

def install(package):
    """
    Automatically runs the Python pip install command to download necessary external packages

    Parameter: package (str) - package to install
    """
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    from pywinauto import ElementNotFoundError
    from pywinauto.application import Application
    from pywinauto.controls.menuwrapper import MenuItemNotEnabled
    from pathlib import Path
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
except:
    install("pywinauto")
    install("pandas")
    install("openpyxl")

    # Reimport after installation
    from pywinauto import ElementNotFoundError
    from pywinauto.application import Application
    from pywinauto.controls.menuwrapper import MenuItemNotEnabled
    from pathlib import Path
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    tkinter.messagebox.showinfo("ERROR", "Python modules could not be imported. Check terminal output. Program has terminated.")
    print("ERROR: Python modules could not be imported.")
    sys.exit(1)

def get_txt_file(directory, message):
    """
    Opens a window to select a text file

    Parameters:
    directory (str) - folder to open
    message (str) - title for window

    Return:
    (str) - file path
    """
    root = tkinter.Tk()
    # Hide the main window
    root.withdraw() 
    root.update()
    file = tkinter.filedialog.askopenfilename(title=message, initialdir=directory, filetypes=[("Text files", "*.TXT *.txt")])
    root.destroy()
    return file

def access_ASPEN():
    """
    Connects to open ASPEN Oneliner window to obtain TTY Window data

    Returns:
    folder_path (Path) - directory of txt file saved
    distance_curve (bool) - whether the curve examined is Distance or Overcurrent 
    """
    timeout_seconds = 360
    try:
        start_time = time.time()
        app = None
        # Connect to ASPEN Oneliner
        # Requires ASPEN Oneliner to already be running
        while time.time() - start_time < timeout_seconds:
            try:
                app = Application(backend="win32").connect(title_re=".*ASPEN OneLiner.*")
                print("Connected to ASPEN OneLiner.")
                break
            except Exception:
                time.sleep(0.5)
        if not app:
            raise ElementNotFoundError()
        # Access main window
        main_window = app.window(title_re=".*ASPEN OneLiner.*")
        
        time.sleep(0.2)

        # Safely close windows before accessing
        # Assumes the title of windows contains phrases below
        distance_curve_window = app.window(title_re=".*Distance.*")
        overcurrent_curve_window = app.window(title_re=".*Overcurrent.*")
        temp_window = app.window(title_re=".*Show Relay Curve.*")
        tty_window = app.window(title_re=".*TTY.*")
        if distance_curve_window.exists():
            distance_curve_window.close()
        if overcurrent_curve_window.exists():
            overcurrent_curve_window.close()
        if temp_window.exists():
            temp_window.close()
        if tty_window.exists():
            tty_window.close()

        # Show Relay Curve and Logic Scheme
        main_window.menu_select("Relay->View Relay Curve and Logic Scheme...")

        start_time = time.time()
        
        # Wait for the user to select a relay and click OK
        print("Waiting for user input...")
        try:
            while True:
                if distance_curve_window.exists(timeout=1):
                    distance_curve_window.wait("visible", timeout=10)
                    print("Relay curve window detected.")
                    break
                elif overcurrent_curve_window.exists(timeout=1):
                    overcurrent_curve_window.wait("visible", timeout=10)
                    print("Relay curve window detected.")
                    break
                elif time.time() - start_time > timeout_seconds:
                    print("Timeout: No relay curve window opened within 360 seconds.")
                    tkinter.messagebox.showerror("Timeout", "Relay Curve window did not open. Program terminated.")
                    sys.exit(1)
        except Exception as e:
            print(f"ERROR: {e}")
            tkinter.messagebox.showerror("ERROR", "An error occurred. Program terminated.")
            sys.exit(1)
        
        # Use curve window
        # Show Relay Operations for All Faults
        if distance_curve_window.exists():
            distance_curve = True
            distance_curve_window.set_focus()
            distance_curve_window.menu_select("Show->Relay Operations for All Faults")
        elif overcurrent_curve_window.exists():
            distance_curve = False
            overcurrent_curve_window.set_focus()
            overcurrent_curve_window.menu_select("Show->Relay Operations for All Faults")
        
        print("Relay curve window ready.")
        print("Showing relay operations for all faults.")

        start_time = time.time()

        while time.time() - start_time < timeout_seconds:
            try:
                # Open TTY window
                main_window.menu_select("View->TTY Window")
                break
            except Exception:
                time.sleep(0.5)
        if not app:
            raise TimeoutError()
        # Select all text in TTY Window
        tty_window.menu_select("Edit->Select All")
        # Save all text to txt file
        tty_window.menu_select("TTY->Save Selected Text...")
        
        # Wait for File Explorer to open
        file_dialog = None
        try:
            file_dialog = app.window(title_re=".*Write Selected Text To.*")
            file_dialog.wait("visible", timeout=timeout_seconds)
        except Exception:
            print("Timeout: File Save dialog never appeared.")
            tkinter.messagebox.showerror("Timeout", " File Save dialog never appeared. Program terminated.")
            sys.exit(1)

        # Wait for user to save file
        print("Waiting for user input...")
        folder_path = ""
        
        # Extract file folder
        while file_dialog.exists() and (time.time() - start_time) < timeout_seconds:
            try:
                # Address bar
                address_bar = file_dialog.child_window(title_re="Address:.*", class_name="ToolbarWindow32")
                # Get folder path
                folder_texts = address_bar.texts()
                if folder_texts:
                    folder_path = folder_texts[0].strip()
            except Exception:
                pass
            time.sleep(0.1)

        print(f"Detected folder path: {folder_path}")

        # Wait for dialog to close
        try:
            file_dialog.wait_not("visible", timeout=timeout_seconds)
        except Exception:
            print("Timeout: File dialog did not close.")
            tkinter.messagebox.showerror("Timeout", "File dialog did not close. Program terminated.")
            sys.exit(1)

        # Return TTY Window directory & the type of curve
        if folder_path != "":
            tty_window.close()
            if distance_curve_window.exists():
                distance_curve_window.close()
            elif overcurrent_curve_window.exists():
                overcurrent_curve_window.close()
            return Path(folder_path), distance_curve
        print("File not saved/selected.")
        tkinter.messagebox.showerror("File Not Found", "Could not find saved file. Program has terminated.")
        sys.exit(0)
    except ElementNotFoundError as e:
        print(f"Error connecting to ASPEN OneLiner: {e}")
        tkinter.messagebox.showerror("ERROR", "Could not connect to ASPEN.\nTry closing & reopening ASPEN OneLiner.\nProgram has terminated.")
        sys.exit(1)
    except MenuItemNotEnabled as e:
        print(f"Error: No fault detected &/or no relay selected: {e}")
        tkinter.messagebox.showerror("ERROR", "No fault detected.\nPlease run fault(s) manually & select a relay before starting program.\nProgram has terminated.")
        sys.exit(1)

def get_fault_descriptions(lines):
    """
    Iterates over Fault Description section of TTY window text file to find lines containing pertinent information.

    Parameter:
    lines (list) - lines of text from the Fault Description section

    Return:
    fault_descriptions (DataFrame) - categorized data parsed from Fault Description section
    """
    # Match Fault lines
    fault_lines = []
    with_end_open = []
    for idx, current_line in enumerate(lines):
        if "on:" not in current_line.lower():
            continue
        current_fault = current_line.strip()
        next_line = lines[idx + 1] if idx + 1 < len(lines) else ""
        if "branch" in next_line.lower() and "outage" in next_line.lower():
            current_outage = next_line.strip()
            fault_lines.append([current_fault, current_outage])
            with_end_open.append(False)
        elif "with end opened" in next_line.lower():
            next_next_line = lines[idx + 2] if idx + 2 < len(lines) else ""
            if "branch" in next_next_line.lower() and "outage" in next_next_line.lower():
                current_outage = next_next_line.strip()
                fault_lines.append([current_fault, current_outage])
            else:
                fault_lines.append([current_fault, ""])
            with_end_open.append(True)
        else:
            fault_lines.append([current_fault, ""])
            with_end_open.append(False)
    # Get fault lines
    fault_line_1 = [line[0] for line in fault_lines]
    # Get Branch Outage line
    fault_line_2 = [line[1]for line in fault_lines]
    # Fault #
    fault_desc_nums = []
    # Fault sim
    fault_sims = []
    # Faulted line
    faulted_lines = []
    # Fault types
    fault_types = []
    # Contingency
    contingencies = []
    # Iterate lines to get fault information
    for line in fault_line_1:
        if line == "":
            continue
        # Get fault #
        current = re.search(r'\b(\d+\.\s)(?!kV)', line)
        if current:
            fault_desc_nums.append(current.group().split(".")[0])
        else:
            fault_desc_nums.append("")
        # Get fault sim
        current = re.search(r'\.\s.*?:', line)
        if current:
            current = current.group()[2:-4]
            if "interm." in current.lower():
                # Get percent on the line for Interm Faults
                percent_line = [phrase for phrase in line.split() if "%" in phrase]
                percent_line = "".join(percent_line)
                percent_line = [percentage for percentage in re.split(r'[()]', percent_line)]
                percent = next((p for p in reversed(percent_line) if "%" in p))
                fault_sims.append(f"{current} {percent}")
            else:
                fault_sims.append(current)
        else:
            fault_sims.append("")
        # Get faulted line
        bus_start = line.rfind("on:") + 3
        bus_end = line.rfind("kV ") + 6
        # Parse fault type
        current = re.search(r'\dLG', line[bus_start:bus_end])
        if current:
            current = current.group()
            faulted_lines.append(re.sub(r'\dLG', '', line[bus_start:bus_end].strip()))
        elif re.search(r'kV\s+LL', line[bus_start:bus_end]):
            current = re.search(r'kV\s+LL', line[bus_start:bus_end]).group()
            faulted_lines.append(re.sub(r'\s+LL\s*', '', line[bus_start:bus_end].strip()))
        else:
            faulted_lines.append(line[bus_start:bus_end].strip())
        current = re.search(r'\dLG', line)
        if not current:
            current = re.search(r'kV\s+LL\s*', line)
            if not current:
                current = re.search(r'L\s+LL\s*', line)
                if not current:
                    current = ""
                else:
                    current = current.group().strip().split()[1]
            else:
                current = current.group().strip().split()[1]
        else:
            current = current.group().strip()
        fault_types.append(current)
    # Iterate over branch outage lines
    for line in fault_line_2:
        if line == "":
            contingencies.append("")
        else:
            current = line[line.lower().rfind("outage:"):].replace("outage:", "").strip()
            contingencies.append(current)

    # Make Dataframe entries for fault description information
    rows_for_desc_frame = []
    for fault_desc_num, fault_sim, end_open, f_line, fault_type, outage in zip(fault_desc_nums, fault_sims, with_end_open, faulted_lines, fault_types, contingencies):
        
        if fault_desc_num != "" and fault_sim != "" and f_line != "" and fault_type != "":
            suffix = "with end opened" if end_open == True else ""
            if end_open:
                suffix = "with end opened"
                if '%' in fault_sim:
                    percent = fault_sim.split()[-1]
                    final_sim = f"{" ".join(fault_sim.split()[:-1])} {suffix} {percent}"
                else:
                    final_sim = f"{fault_sim} {suffix}"
            else:
                final_sim = fault_sim
            
            rows_for_desc_frame.append({
                    "Fault #": int(fault_desc_num),
                    "Fault Sim": final_sim,
                    "Faulted Line": f_line,
                    "Fault Type": fault_type,
                    "Branch Outage": outage
                })
        
    if not rows_for_desc_frame:
        raise ValueError("ERROR: No fault description data could be parsed from the TTY text.")
    fault_descriptions = pd.DataFrame(rows_for_desc_frame)
    return fault_descriptions

def get_fault_table(lines, curve_type):
    """
    Iterate over fault table section to find lines containing pertinent information.

    Parameter:
    lines (list) - lines of text from the fault table section

    Return:
    fault_descriptions (DataFrame) - categorized data parsed from fault table section
    """
    rows_for_table_frame = []
    for i in range(0, len(lines), 4):
        # Check for at least 4 lines left
        if i + 3 >= len(lines):
            break
        fault_line = lines[i]
        relay_line = lines[i+2]
        if curve_type:
            imp_line = lines[i+3]
        else:
            time_line = lines[i+3]

        # print(f"\nParsing block starting at line {i}:")
        # print(f"Line 1: {fault_line}")
        # print(f"Line 3: {lines[i+2]}")
        # print(f"Line 4:   {lines[i+3]}")
        # print()

        # Get relay name
        if curve_type:
            relay_name = relay_line.split(":")[0].split()
        else:
            relay_name = relay_line.split(".")[0].split()
        relay_name = " ".join(relay_name[:-1])
        # Get fault number
        fault_nums = re.findall(r'Fault\s+(\d+)', fault_line)
        # Get time
        if curve_type:
            time_line = relay_line.split(":")
            time_line.insert(1, time_line[0][-3:])
            time_line[0] = time_line[0][:-3]
            time_line = "".join(time_line[1:]).split()
            time_pair = []
            for idx, t in enumerate(time_line):
                if idx + 1 < len(time_line):
                    next_val = time_line[idx + 1]
                    if idx % 2 == 0:
                        time_pair.append([t, next_val])
            # Get impedance
            imp_parts = imp_line.split("@")
            imp_parts = " ".join(imp_parts).split()
            imp_pairs = []
            for idx, imp in enumerate(imp_parts):
                if idx + 1 < len(imp_parts):
                    next_val = imp_parts[idx + 1]
                    if idx % 2 == 0:
                        imp_pairs.append([imp, next_val])
            # Make Distance Curve DataFrame entries
            for fault, time_label, imp in zip(fault_nums, time_pair, imp_pairs):
                
                rows_for_table_frame.append({
                    "Relay": relay_name,
                    "Fault #": int(fault) if fault != "" else "",
                    "Operate Time": time_label[1],
                    "Operate Zone": time_label[0],
                    "Impedance (Magnitude)": float(imp[0]),
                    "Impedance (Angle)": float(imp[1])
                })
        else:
            time_parts = time_line.split()
            # Get fault current
            fault_current_line = relay_line.split()
            fault_current_line = [entry for entry in fault_current_line if "." in entry]

            # Make Overcurrent Curve DataFrame entries
            for fault, time, currentA in zip(fault_nums, time_parts, fault_current_line):
                rows_for_table_frame.append({
                    "Relay": relay_name,
                    "Fault #": int(fault) if fault != "" else "",
                    "Operate Time": time,
                    "Fault Current": currentA,
                })
    if not rows_for_table_frame:
        raise ValueError("ERROR: No relay fault data could be parsed from the TTY text.")
    fault_table = pd.DataFrame(rows_for_table_frame)
    return fault_table

def clean_tty_text(tty_text_path, fault_start, curve_type):
    """
    Reads all content in the TTY window text file. Stores, parses, & organizes last simulated fault data.

    Parameters:
    tty_text_path (str) - file path to TTY window txt file
    fault_start (str) - phrase to look for in txt file to indicate the beginning of the data to store\
    curve_type (bool) - whether the curve examined is Distance or Overcurrent 
    """
    # Read all TTY window text saved
    with open(tty_text_path, 'r', encoding='utf-8') as file:
        content = file.read()

    # Find where fault info begins in txt file
    for phrase in fault_start[::-1]:
        fault_info_start = content.rfind(phrase)
        if fault_info_start == -1:
            raise ValueError(f"ERROR: The string '{phrase}' was not found in the TTY text.")

    # Keep the text starting from the phrase
    new_content = content[fault_info_start:]

    # Write fault info back to the txt file
    with open(tty_text_path, 'w', encoding='utf-8') as file:
        file.write(new_content)
    
    print(f"TTY text file cleaned.")

    # Get fault descriptions
    description_list = new_content[:content.rfind("Fault  1 ")]
    # Separate into list of lines
    lines = [line for line in description_list.strip().splitlines() if line.strip()]

    # Parse fault description section of TTY Window
    fault_description_frame = get_fault_descriptions(lines)
    print(len(fault_description_frame), "Fault descriptions found.")

    # Parse fault table section of TTY window
    fault_table_list = new_content[new_content.rfind("Fault  1 "):]
    lines = [line for line in fault_table_list.strip().splitlines() if line.strip()]
    fault_table_frame = get_fault_table(lines, curve_type)
    print(len(fault_table_frame), "Fault table entries found.")

    # Match & merge DataFrames
    faults_frame = fault_description_frame.merge(fault_table_frame, on="Fault #", how='inner')
    faults_frame.set_index("Fault #")
    return faults_frame

def get_max_impedance(dataframe):
    """
    Gets the maximum impedance from the faults collected

    Parameter:
    dataframe (DataFrame) - table of faults to compare

    Returns:
    max_impedance (float) - maximum overall impedance
    max_impedance_by_relay (list) - list of maximum impedance amongst relays
    """
    # Overall Max
    max_impedance = dataframe['Impedance (Magnitude)'].max()
    # Maximums by relay
    max_impedance_by_relay = dataframe.groupby('Relay')['Impedance (Magnitude)'].max()
    max_impedance_by_relay = max_impedance_by_relay.to_dict()
    return max_impedance, max_impedance_by_relay

def get_min_impedance(dataframe):
    """
    Gets the minimum impedance from the faults collected

    Parameter:
    dataframe (DataFrame) - table of faults to compare

    Returns:
    min_impedance (float) - minimum overall impedance
    min_impedance_by_relay (list) - list of minimum impedance amongst relays
    """
    # Overall Min
    min_impedance = dataframe['Impedance (Magnitude)'].min()
    # Minimums by relay
    min_impedance_by_relay = dataframe.groupby('Relay')['Impedance (Magnitude)'].min()
    min_impedance_by_relay = min_impedance_by_relay.to_dict()
    return min_impedance, min_impedance_by_relay

def apply_header_style(cell):
    """
    Applies appearance for headers

    Parameter:
    cell - cell in Excel file to apply header appearance to
    """
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.font = header_font
    cell.fill = header_fill

def apply_row_style(cell):
    """
    Applies appearance for row cell

    Parameter:
    cell - cell in Excel file to apply row appearance to
    """
    fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    cell.fill = fill_gray

def stylize_main_table(ws, end_col):
    """
    Applies appearance to main table in Excel file

    Parameters:
    ws (Worksheet) - sheet of Excel file to use style on
    end_col (int) - index of last column of main table in Excel file
    """
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Style header row
    for cell in ws[1]:
        if cell.column > end_col:
            break
        apply_header_style(cell)
        cell.border = thin_border

    # Style alternating data rows for banded rows effect
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for col_idx, cell in enumerate(row, start=1):
            if col_idx > end_col:
                break
            # Apply gray fill to even rows
            if row_idx % 2 == 0:
                apply_row_style(cell)
            # Apply borders and alignment to all data cells
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

def create_xlsx(txt_location, dataframe):
    """
    Creates or appends to an Excel spreadsheet to display fault data.

    Parameters:
    txt_location (str) - file path of TTY Window txt file
    dataframe (DataFrame) - fault data
    """
    # Create Excel file
    original_txt_location = Path(txt_location)
    file_path = original_txt_location.with_name("Fault Summary - TTY Window.xlsx")
    sheet_name = "Fault Summary"

    if file_path.exists():
        writer_args = {
            'engine': 'openpyxl',
            'mode': 'a',
            'if_sheet_exists': 'replace'
        }
    else:
        writer_args = {
            'engine': 'openpyxl',
            'mode': 'w'
        }

    # Convert DataFrame to Excel spreadsheet
    while True:
        try:
            with pd.ExcelWriter(file_path, **writer_args) as writer:
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
                break
        except PermissionError as e:
            print("Permission Error:", e)
            tkinter.messagebox.showinfo("ERROR", "Please close any already open Summary spreadsheets before program can continue.")
    
    # Load workbook + worksheet for stylization
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    if "Impedance (Magnitude)" in dataframe.columns:
        start_row = 2
        # Impedance maximums
        # Per relay
        max_impedance, max_imp_by_relay = get_max_impedance(dataframe)
        relays = list(max_imp_by_relay.keys())
        maximum_table_column = get_column_letter(ws.max_column + 3)
        ws[f"{maximum_table_column}{start_row - 1}"].value = "Max Impedance by Relay"
        for relay, imp in max_imp_by_relay.items():
            ws[f"{maximum_table_column}{start_row + relays.index(relay)}"].value = relay
            ws[f"{get_column_letter(column_index_from_string(maximum_table_column) + 1)}{start_row + relays.index(relay)}"].value = imp
            if start_row + relays.index(relay) % 2 != 0:
                apply_row_style(ws[f"{maximum_table_column}{start_row + relays.index(relay)}"])
                apply_row_style(ws[f"{get_column_letter(column_index_from_string(maximum_table_column) + 1)}{start_row + relays.index(relay)}"])
        # Overall max
        ws[f"{maximum_table_column}{start_row + len(max_imp_by_relay) + 2}"].value = "Overall Maximum Impedance"
        ws[f"{get_column_letter(column_index_from_string(maximum_table_column) + 1)}{start_row + len(max_imp_by_relay) + 2}"].value = max_impedance
        apply_row_style(ws[f"{get_column_letter(column_index_from_string(maximum_table_column) + 1)}{start_row + len(max_imp_by_relay) + 2}"])

        # Impedance minimums
        # Per relay
        min_impedance, min_imp_by_relay = get_min_impedance(dataframe)
        minimum_table_column = get_column_letter(ws.max_column + 3)
        ws[f"{minimum_table_column}{start_row - 1}"].value = "Min Impedance by Relay"
        for relay, imp in min_imp_by_relay.items():
            ws[f"{minimum_table_column}{start_row + relays.index(relay)}"].value = relay
            ws[f"{get_column_letter(column_index_from_string(minimum_table_column) + 1)}{start_row + relays.index(relay)}"].value = imp
            if start_row + relays.index(relay) % 2 != 0:
                apply_row_style(ws[f"{minimum_table_column}{start_row + relays.index(relay)}"])
                apply_row_style(ws[f"{get_column_letter(column_index_from_string(minimum_table_column) + 1)}{start_row + relays.index(relay)}"])
        # Overall min
        ws[f"{minimum_table_column}{start_row + len(min_imp_by_relay) + 2}"].value = "Overall Minimum Impedance"
        ws[f"{get_column_letter(column_index_from_string(minimum_table_column) + 1)}{start_row + len(min_imp_by_relay) + 2}"].value = min_impedance
        apply_row_style(ws[f"{get_column_letter(column_index_from_string(minimum_table_column) + 1)}{start_row + len(min_imp_by_relay) + 2}"])
        # Stylize max
        apply_header_style(ws[f"{maximum_table_column}{start_row - 1}"])
        ws.merge_cells(f"{maximum_table_column}{start_row - 1}:{get_column_letter(column_index_from_string(maximum_table_column) + 1)}{start_row - 1}")
        apply_header_style(ws[f"{maximum_table_column}{start_row + len(max_imp_by_relay) + 2}"])
        # Stylize min
        apply_header_style(ws[f"{minimum_table_column}{start_row - 1}"])
        ws.merge_cells(f"{minimum_table_column}{start_row - 1}:{get_column_letter(column_index_from_string(minimum_table_column) + 1)}{start_row - 1}")
        apply_header_style(ws[f"{minimum_table_column}{start_row + len(min_imp_by_relay) + 2}"])
    # Stylize main table
    stylize_main_table(ws, len(dataframe.columns))

    # Save workbook
    wb.save(file_path)
    return file_path

def main():
    tty_folder_path, curve_type = access_ASPEN()
    tty_path = get_txt_file(tty_folder_path, "Please select your TTY window output file.")
    print(f"Text file location: {tty_path}")
    faults = clean_tty_text(tty_path, ["Fault description:"], curve_type)
    print("Spreadsheet saved at:", create_xlsx(tty_path, faults))
    return 0

if __name__ == "__main__":
    main()